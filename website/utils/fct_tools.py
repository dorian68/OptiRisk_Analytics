from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.http import JsonResponse, HttpResponse
import pandas as pd

def style_excel_report(excel_file, sheet_name="TableauFormaté"):
    """Personnalise l'apparence de l'Excel avant export HTML"""
    
    # Charger le fichier Excel
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Définition des styles
    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Appliquer le style aux en-têtes (première ligne)
    for cell in ws[1]:
        cell.font = title_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuster la largeur des colonnes automatiquement
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtenir la lettre de la colonne
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustement

    # Sauvegarder les modifications
    wb.save(excel_file)
    print(f"📊 Fichier Excel stylisé: {excel_file}")

def export_excel_to_html_with_css(excel_file, sheet_name="TableauFormaté", output_html="tableau.html"):
    """Exporte un tableau Excel en HTML avec un style CSS"""
    
    # Charger la feuille Excel
    df = pd.read_excel(excel_file, sheet_name=sheet_name)

    # Définition du style CSS
    css_style = """
    <style>
        table { width: 100%; border-collapse: collapse; }
        th { background-color: #4F81BD; color: white; padding: 10px; text-align: center; }
        td { padding: 8px; text-align: center; border: 1px solid #ddd; }
        tr:nth-child(even) { background-color: #f2f2f2; }
    </style>
    """

    # Convertir en HTML avec le CSS
    html_table = df.to_html(index=False, border=0, justify="center")

    # Ajouter le CSS en haut du fichier
    full_html = f"{css_style}\n{html_table}"

    # Sauvegarder le fichier HTML
    with open(output_html, "w", encoding="utf-8") as f:
        f.write(full_html)

    print(f"📄 Rapport exporté en HTML avec CSS: {output_html}")

    return output_html
 
def process_csv(request):
    if request.method == 'POST' and request.FILES.get('file'):
        try:
            # Fichier CSV envoyé
            file = request.FILES['file']

            # Fichiers temporaires
            excel_file = "/tmp/fichier_stylisé.xlsx"
            output_html = "/tmp/tableau_stylisé.html"

            # Étapes :
            import_csv_to_excel(file, excel_file)  # Charger le CSV
            style_excel_report(excel_file)  # Appliquer la mise en page Excel
            html_path = export_excel_to_html_with_css(excel_file)  # Générer le HTML

            # Lire le HTML
            with open(html_path, "r", encoding="utf-8") as f:
                html_content = f.read()

            return HttpResponse(html_content, content_type="text/html")

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Aucun fichier fourni"}, status=400)

def import_csv_to_excel(csv_file, excel_file):
    """Convertit un fichier CSV en fichier Excel"""
    df = pd.read_csv(csv_file)  # Charger le CSV en DataFrame
    df.to_excel(excel_file, index=False, engine="openpyxl")  # Sauvegarder en Excel
    print(f"📂 CSV converti en Excel: {excel_file}")
